#!/usr/bin/env python3
"""Raw STFT / time-frequency audit for the ferron materials bridge.

This module inspects the downloaded Figshare XLSX files for actual
time-frequency amplitude grids. It is intentionally conservative: STFT-labeled
time-only traces are reported as candidates but are not treated as raw
time-frequency maps unless both time and frequency axes are present.
"""

from __future__ import annotations

import csv
import json
import math
import re
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

from spectral_feature_audit import detect_target_peak, estimate_peak_to_background


ROOT = Path(__file__).resolve().parent
OUTPUTS_DIR = ROOT / "outputs"
RAW_DIR = OUTPUTS_DIR / "raw"

TARGET_FREQUENCY_THZ = 3.13
SEARCH_WINDOW_THZ = 0.20
COLLAPSE_THRESHOLD = 0.70
SUSTAIN_STEPS = 2

SUMMARY_FIELDNAMES = [
    "source_file",
    "sheet_name",
    "map_id",
    "time_ps",
    "target_frequency_THz",
    "peak_found",
    "peak_frequency_THz",
    "peak_amplitude",
    "peak_to_background",
    "frequency_drift_from_target_THz",
    "amplitude_retention",
    "frequency_stability",
    "target_band_recoverability",
    "delta_persistence",
    "k_star",
    "safety_margin",
    "confidence",
    "visible_failure",
]

GROUP_FIELDNAMES = [
    "map_id",
    "n_time_points",
    "mean_peak_frequency_THz",
    "std_peak_frequency_THz",
    "mean_peak_amplitude",
    "mean_target_band_recoverability",
    "min_target_band_recoverability",
    "max_target_band_recoverability",
    "k_star_detected",
    "k_star_time_ps",
    "visible_failure_detected",
    "first_visible_failure_time_ps",
]

STALE_STFT_OUTPUTS = [
    "stft_feature_summary.csv",
    "stft_recoverability_by_time.csv",
    "stft_target_band_recoverability.png",
    "stft_peak_frequency_over_time.png",
    "stft_amplitude_map_examples.png",
    "stft_delta_persistence.png",
    "stft_safety_margin.png",
]

TIME_ALIASES = {
    "time",
    "timeps",
    "delay",
    "delayps",
    "dt",
    "dtp",
    "dtps",
    "pumpprobedelay",
    "pumpprobedelayps",
}
FREQUENCY_ALIASES = {"frequency", "frequencythz", "freq", "freqthz", "thz"}
AMPLITUDE_ALIASES = {
    "amplitude",
    "intensity",
    "power",
    "signal",
    "stftamplitude",
    "stftint",
    "spectrogram",
    "map",
    "fourieramplitude",
}
STFT_KEYWORDS = {"stft", "spectrogram", "shorttime", "short_time", "timefrequency", "time_frequency"}


def inspect_possible_stft_sheets(
    root: Path = ROOT,
    *,
    raw_dir: Path | None = None,
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
    search_window_THz: float = SEARCH_WINDOW_THZ,
) -> dict[str, Any]:
    """Scan XLSX sheets and identify possible raw STFT/time-frequency tables."""

    del target_frequency_THz, search_window_THz
    raw_path = raw_dir or root / "outputs" / "raw"
    raw_files = sorted(raw_path.glob("*.xlsx"))
    sheet_reports: list[dict[str, Any]] = []
    file_reports = [
        {
            "file_name": path.name,
            "relative_path": _relative_to_root(path, root),
            "size_bytes": path.stat().st_size,
        }
        for path in raw_files
        if path.is_file()
    ]

    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        return {
            "openpyxl_available": False,
            "raw_files_inspected": file_reports,
            "candidate_sheets_inspected": 0,
            "candidate_sheet_reports": [],
            "sheet_reports": [],
            "missing_data_notes": [f"openpyxl unavailable for XLSX STFT inspection: {exc}"],
        }

    for path in raw_files:
        if not path.is_file():
            continue
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                preview_rows = list(sheet.iter_rows(values_only=True, max_row=12))
                sheet_reports.append(
                    _inspect_sheet_preview(
                        path,
                        sheet.title,
                        preview_rows,
                        max_row=sheet.max_row,
                        max_column=sheet.max_column,
                        root=root,
                    )
                )
        finally:
            workbook.close()

    candidate_reports = [report for report in sheet_reports if report.get("candidate")]
    return {
        "openpyxl_available": True,
        "raw_files_inspected": file_reports,
        "candidate_sheets_inspected": len(candidate_reports),
        "candidate_sheet_reports": candidate_reports,
        "sheet_reports": sheet_reports,
        "missing_data_notes": [],
    }


def parse_stft_long_form(
    raw_rows: list[tuple[Any, ...]],
    *,
    source_file: str,
    sheet_name: str,
) -> tuple[dict[str, Any] | None, str | None]:
    """Parse Shape A: long-form time/frequency/amplitude rows."""

    header_index, headers = _find_header_row(raw_rows)
    if header_index is None or headers is None:
        return None, "no_header_row_detected"

    normalized = [_normalize_header(value, index) for index, value in enumerate(headers)]
    time_index = _find_header_index(normalized, TIME_ALIASES)
    frequency_index = _find_header_index(normalized, FREQUENCY_ALIASES)
    amplitude_index = _find_header_index(normalized, AMPLITUDE_ALIASES)
    if time_index is None:
        return None, "long_form_time_column_missing"
    if frequency_index is None:
        return None, "long_form_frequency_column_missing"
    if amplitude_index is None:
        return None, "long_form_amplitude_column_missing"

    time_values: list[float] = []
    frequency_values: list[float] = []
    amplitude_values: list[float] = []
    for raw_row in raw_rows[header_index + 1 :]:
        time_value = _float_or_none(_cell(raw_row, time_index))
        frequency_value = _float_or_none(_cell(raw_row, frequency_index))
        amplitude_value = _float_or_none(_cell(raw_row, amplitude_index))
        if time_value is None or frequency_value is None or amplitude_value is None:
            continue
        time_values.append(time_value)
        frequency_values.append(frequency_value)
        amplitude_values.append(abs(amplitude_value))

    if len(time_values) < 9:
        return None, "long_form_not_enough_numeric_rows"

    stft_map = {
        "source_file": source_file,
        "sheet_name": sheet_name,
        "map_id": _safe_id(f"{Path(source_file).stem}_{sheet_name}_long_form"),
        "parse_shape": "long_form",
        "time_ps": np.asarray(time_values, dtype=float),
        "frequency_THz": np.asarray(frequency_values, dtype=float),
        "amplitude": np.asarray(amplitude_values, dtype=float),
    }
    return stft_map, None


def parse_stft_matrix_form(
    raw_rows: list[tuple[Any, ...]],
    *,
    source_file: str,
    sheet_name: str,
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
    search_window_THz: float = SEARCH_WINDOW_THZ,
) -> tuple[dict[str, Any] | None, str | None]:
    """Parse Shape B: a 2D amplitude matrix with numeric time/frequency axes."""

    if len(raw_rows) < 4:
        return None, "matrix_not_enough_rows"
    header = list(raw_rows[0])
    if len(header) < 4:
        return None, "matrix_not_enough_columns"

    column_axis = [_float_or_none(value) for value in header[1:]]
    column_axis_pairs = [(index + 1, value) for index, value in enumerate(column_axis) if value is not None]
    row_axis_pairs: list[tuple[int, float]] = []
    for row_index, raw_row in enumerate(raw_rows[1:], start=1):
        value = _float_or_none(_cell(raw_row, 0))
        if value is not None:
            row_axis_pairs.append((row_index, value))

    if len(column_axis_pairs) < 3 or len(row_axis_pairs) < 3:
        return None, "matrix_axes_not_numeric_enough"

    top_left = _text(_cell(raw_rows[0], 0))
    sheet_text = _clean_key(f"{sheet_name} {top_left} {' '.join(_text(value) for value in header[:5])}")
    column_values = [value for _, value in column_axis_pairs]
    row_values = [value for _, value in row_axis_pairs]
    column_is_frequency = _axis_includes_target(column_values, target_frequency_THz, search_window_THz)
    row_is_frequency = _axis_includes_target(row_values, target_frequency_THz, search_window_THz)
    has_time_hint = any(token in sheet_text for token in ("time", "delay", "dt", "ps"))
    has_frequency_hint = any(token in sheet_text for token in ("frequency", "freq", "thz"))

    if column_is_frequency and not row_is_frequency:
        time_axis = row_axis_pairs
        frequency_axis = column_axis_pairs
        orientation = "rows_time_columns_frequency"
    elif row_is_frequency and not column_is_frequency:
        time_axis = column_axis_pairs
        frequency_axis = row_axis_pairs
        orientation = "rows_frequency_columns_time"
    elif column_is_frequency and row_is_frequency:
        return None, "matrix_both_axes_overlap_target_frequency"
    elif not has_time_hint or not has_frequency_hint:
        return None, "matrix_axes_not_identifiable_as_time_and_frequency"
    else:
        return None, "matrix_frequency_axis_does_not_include_target_window"

    time_values: list[float] = []
    frequency_values: list[float] = []
    amplitude_values: list[float] = []
    if orientation == "rows_time_columns_frequency":
        for row_index, time_value in time_axis:
            raw_row = raw_rows[row_index]
            for column_index, frequency_value in frequency_axis:
                amplitude_value = _float_or_none(_cell(raw_row, column_index))
                if amplitude_value is None:
                    continue
                time_values.append(time_value)
                frequency_values.append(frequency_value)
                amplitude_values.append(abs(amplitude_value))
    else:
        for row_index, frequency_value in frequency_axis:
            raw_row = raw_rows[row_index]
            for column_index, time_value in time_axis:
                amplitude_value = _float_or_none(_cell(raw_row, column_index))
                if amplitude_value is None:
                    continue
                time_values.append(time_value)
                frequency_values.append(frequency_value)
                amplitude_values.append(abs(amplitude_value))

    if len(time_values) < 9:
        return None, "matrix_not_enough_finite_amplitudes"
    return (
        {
            "source_file": source_file,
            "sheet_name": sheet_name,
            "map_id": _safe_id(f"{Path(source_file).stem}_{sheet_name}_matrix"),
            "parse_shape": f"matrix_{orientation}",
            "time_ps": np.asarray(time_values, dtype=float),
            "frequency_THz": np.asarray(frequency_values, dtype=float),
            "amplitude": np.asarray(amplitude_values, dtype=float),
        },
        None,
    )


def validate_stft_grid(
    stft_map: dict[str, Any],
    *,
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
    search_window_THz: float = SEARCH_WINDOW_THZ,
) -> tuple[bool, str | None]:
    """Confirm that a parsed STFT map has usable time, frequency, and amplitude axes."""

    time = np.asarray(stft_map.get("time_ps"), dtype=float)
    frequency = np.asarray(stft_map.get("frequency_THz"), dtype=float)
    amplitude = np.asarray(stft_map.get("amplitude"), dtype=float)
    finite = np.isfinite(time) & np.isfinite(frequency) & np.isfinite(amplitude)
    if not np.any(finite):
        return False, "stft_grid_has_no_finite_values"
    time = time[finite]
    frequency = frequency[finite]
    amplitude = amplitude[finite]
    unique_time = np.unique(time)
    unique_frequency = np.unique(frequency)
    if unique_time.size < 3:
        return False, "stft_grid_requires_at_least_3_time_points"
    if unique_frequency.size < 3:
        return False, "stft_grid_requires_at_least_3_frequency_points"
    target_window = np.abs(unique_frequency - target_frequency_THz) <= search_window_THz
    if not np.any(target_window):
        return False, "frequency_axis_does_not_cover_target_window"
    if amplitude.size < 9:
        return False, "stft_grid_requires_at_least_9_finite_amplitudes"
    return True, None


def extract_target_band_over_time(
    stft_map: dict[str, Any],
    *,
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
    search_window_THz: float = SEARCH_WINDOW_THZ,
) -> list[dict[str, Any]]:
    """Extract target-window peak records for each time slice."""

    time = np.asarray(stft_map.get("time_ps"), dtype=float)
    frequency = np.asarray(stft_map.get("frequency_THz"), dtype=float)
    amplitude = np.asarray(stft_map.get("amplitude"), dtype=float)
    finite = np.isfinite(time) & np.isfinite(frequency) & np.isfinite(amplitude)
    time = time[finite]
    frequency = frequency[finite]
    amplitude = np.abs(amplitude[finite])
    rows: list[dict[str, Any]] = []
    for time_value in sorted(np.unique(time)):
        mask = time == time_value
        if np.count_nonzero(mask) < 3:
            continue
        freq_slice = frequency[mask]
        amp_slice = amplitude[mask]
        order = np.argsort(freq_slice)
        freq_slice = freq_slice[order]
        amp_slice = amp_slice[order]
        peak = detect_target_peak(
            freq_slice,
            amp_slice,
            target_frequency_THz=target_frequency_THz,
            search_window_THz=search_window_THz,
        )
        peak_frequency = _float_or_none(peak.get("peak_frequency_THz"))
        peak_amplitude = _float_or_none(peak.get("peak_amplitude"))
        peak_to_background, background_note = estimate_peak_to_background(
            freq_slice,
            amp_slice,
            peak_amplitude,
            target_frequency_THz=target_frequency_THz,
            search_window_THz=search_window_THz,
        )
        rows.append(
            {
                "source_file": stft_map.get("source_file"),
                "sheet_name": stft_map.get("sheet_name"),
                "map_id": stft_map.get("map_id"),
                "time_ps": float(time_value),
                "target_frequency_THz": target_frequency_THz,
                "peak_found": bool(peak.get("peak_found")),
                "peak_frequency_THz": peak_frequency,
                "peak_amplitude": peak_amplitude,
                "peak_to_background": peak_to_background,
                "frequency_drift_from_target_THz": None
                if peak_frequency is None
                else peak_frequency - target_frequency_THz,
                "_background_note": background_note,
            }
        )
    return rows


def compute_stft_recoverability(
    rows: list[dict[str, Any]],
    *,
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
    search_window_THz: float = SEARCH_WINDOW_THZ,
) -> None:
    """Populate amplitude, frequency, background, and combined recoverability proxies."""

    grouped = _group_rows_by_map(rows)
    for map_rows in grouped.values():
        ordered = sorted(map_rows, key=lambda row: _float_or_none(row.get("time_ps")) or 0.0)
        baseline = next((row for row in ordered if row.get("peak_found")), None)
        baseline_amplitude = _float_or_none(baseline.get("peak_amplitude")) if baseline else None
        baseline_p2b = _float_or_none(baseline.get("peak_to_background")) if baseline else None
        previous_found = True
        for row in ordered:
            peak_found = bool(row.get("peak_found"))
            peak_amplitude = _float_or_none(row.get("peak_amplitude"))
            peak_frequency = _float_or_none(row.get("peak_frequency_THz"))
            peak_to_background = _float_or_none(row.get("peak_to_background"))
            amplitude_retention = None
            if baseline_amplitude is not None and baseline_amplitude > 1.0e-12 and peak_amplitude is not None:
                amplitude_retention = _clip01(peak_amplitude / baseline_amplitude)
            frequency_stability = None
            if peak_frequency is not None:
                frequency_stability = _clip01(1.0 - abs(peak_frequency - target_frequency_THz) / search_window_THz)
            p2b_stability = None
            if baseline_p2b is not None and baseline_p2b > 1.0e-12 and peak_to_background is not None:
                p2b_stability = _clip01(peak_to_background / baseline_p2b)
            continuity = 1.0 if peak_found and previous_found else 0.0
            components = [
                amplitude_retention,
                frequency_stability,
                p2b_stability,
                continuity,
            ]
            usable = [value for value in components if value is not None]
            row["amplitude_retention"] = amplitude_retention
            row["frequency_stability"] = frequency_stability
            row["_peak_to_background_stability"] = p2b_stability
            row["target_band_recoverability"] = float(sum(usable) / len(usable)) if usable else None
            row["confidence"] = _confidence_from_background(peak_to_background, row.get("target_band_recoverability"))
            row["visible_failure"] = _visible_failure(row)
            previous_found = peak_found


def compute_stft_delta_persistence(rows: list[dict[str, Any]]) -> None:
    """Populate delta persistence between successive time slices."""

    grouped = _group_rows_by_map(rows)
    for map_rows in grouped.values():
        ordered = sorted(map_rows, key=lambda row: _float_or_none(row.get("time_ps")) or 0.0)
        previous: float | None = None
        for row in ordered:
            current = _float_or_none(row.get("target_band_recoverability"))
            if current is None:
                row["delta_persistence"] = None
            elif previous is None:
                row["delta_persistence"] = 0.0
            else:
                row["delta_persistence"] = current - previous
            if current is not None:
                previous = current


def detect_stft_k_star(
    rows: list[dict[str, Any]],
    *,
    collapse_threshold: float = COLLAPSE_THRESHOLD,
    sustain_steps: int = SUSTAIN_STEPS,
) -> dict[str, float | None]:
    """Detect first sustained target-band recoverability collapse per map."""

    k_star_by_map: dict[str, float | None] = {}
    grouped = _group_rows_by_map(rows)
    for map_id, map_rows in grouped.items():
        ordered = sorted(map_rows, key=lambda row: _float_or_none(row.get("time_ps")) or 0.0)
        k_star_time: float | None = None
        values = [
            (_float_or_none(row.get("target_band_recoverability")), _float_or_none(row.get("time_ps")))
            for row in ordered
        ]
        clean = [(value, time_value) for value, time_value in values if value is not None and time_value is not None]
        for index in range(len(clean)):
            window = clean[index : index + sustain_steps]
            if len(window) < sustain_steps:
                continue
            if all(value < collapse_threshold for value, _ in window):
                k_star_time = window[0][1]
                break
        k_star_by_map[map_id] = k_star_time
        for row in ordered:
            row["k_star"] = k_star_time
            time_value = _float_or_none(row.get("time_ps"))
            if k_star_time is None or time_value is None:
                row["safety_margin"] = None
            elif time_value <= k_star_time:
                row["safety_margin"] = k_star_time - time_value
            else:
                row["safety_margin"] = 0.0
    return k_star_by_map


def summarize_stft_features(
    *,
    stft_maps: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    inspection: dict[str, Any],
    k_star_by_map: dict[str, float | None],
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
) -> tuple[list[str], str, list[str]]:
    """Return bounded plain-English STFT summary lines and interpretation."""

    missing_notes: list[str] = []
    rejected = inspection.get("rejected_candidate_sheets") or []
    for item in rejected:
        missing_notes.append(
            f"{item.get('source_file')}#{item.get('sheet_name')} rejected: {item.get('reason')}"
        )

    if not stft_maps:
        return (
            [],
            "No raw STFT/time-frequency map was parsed. Existing spectral audit result remains unchanged.",
            missing_notes,
        )

    found = [row for row in rows if row.get("peak_found")]
    recoverability = _finite_values(row.get("target_band_recoverability") for row in rows)
    peak_frequencies = _finite_values(row.get("peak_frequency_THz") for row in found)
    amplitudes = _finite_values(row.get("amplitude_retention") for row in found)
    features = [
        f"Usable raw STFT/time-frequency maps parsed: {len(stft_maps)}.",
        f"Target-window peak records found in {len(found)}/{len(rows)} audited time slices near {target_frequency_THz:.2f} THz.",
    ]
    if peak_frequencies:
        features.append(
            "STFT target peak frequency range is "
            f"{min(peak_frequencies):.6g}-{max(peak_frequencies):.6g} THz "
            f"(mean {sum(peak_frequencies) / len(peak_frequencies):.6g} THz)."
        )
    if amplitudes:
        features.append(
            "STFT target-band amplitude retention ranges from "
            f"{min(amplitudes):.6g} to {max(amplitudes):.6g}."
        )
    if recoverability:
        features.append(
            "STFT recoverability ranges from "
            f"{min(recoverability):.6g} to {max(recoverability):.6g}."
        )
    k_star_values = [value for value in k_star_by_map.values() if value is not None]
    if k_star_values:
        interpretation = (
            "A raw STFT/time-frequency map was parsed and the proxy audit detects a sustained "
            f"target-band recoverability crossing beginning at {min(k_star_values):.6g} ps. "
            "This remains a bounded signal-quality result, not a HAOS-IIP proof."
        )
    else:
        interpretation = (
            "Raw STFT/time-frequency data were parsed and the target-window feature remains "
            "recoverable over the audited time slices under the current proxy thresholds; no "
            "STFT k-star collapse is detected."
        )
    return features, interpretation, missing_notes


def run_stft_feature_audit(
    *,
    root: Path = ROOT,
    output_dir: Path = OUTPUTS_DIR,
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
    search_window_THz: float = SEARCH_WINDOW_THZ,
) -> dict[str, Any]:
    """Run the raw XLSX STFT/time-frequency inspection and audit."""

    output_dir.mkdir(parents=True, exist_ok=True)
    inspection = inspect_possible_stft_sheets(
        root,
        target_frequency_THz=target_frequency_THz,
        search_window_THz=search_window_THz,
    )
    accepted_maps: list[dict[str, Any]] = []
    rejected_candidate_sheets: list[dict[str, Any]] = []

    if not inspection.get("openpyxl_available", False):
        summary = _no_stft_summary(inspection, "XLSX parsing requires openpyxl.")
        _remove_stale_stft_outputs(output_dir)
        _write_json(output_dir / "stft_feature_summary.json", summary)
        return summary

    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError:
        summary = _no_stft_summary(inspection, "XLSX parsing requires openpyxl.")
        _remove_stale_stft_outputs(output_dir)
        _write_json(output_dir / "stft_feature_summary.json", summary)
        return summary

    for candidate in inspection.get("candidate_sheet_reports", []):
        relative_path = str(candidate.get("relative_path", ""))
        sheet_name = str(candidate.get("sheet_name", ""))
        path = root / relative_path
        if not path.exists():
            rejected_candidate_sheets.append(_rejection(candidate, "candidate_file_missing"))
            continue
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                rejected_candidate_sheets.append(_rejection(candidate, "candidate_sheet_missing"))
                continue
            raw_rows = list(workbook[sheet_name].iter_rows(values_only=True))
        finally:
            workbook.close()

        stft_map, reason = parse_stft_long_form(
            raw_rows,
            source_file=relative_path,
            sheet_name=sheet_name,
        )
        if stft_map is None:
            matrix_map, matrix_reason = parse_stft_matrix_form(
                raw_rows,
                source_file=relative_path,
                sheet_name=sheet_name,
                target_frequency_THz=target_frequency_THz,
                search_window_THz=search_window_THz,
            )
            stft_map = matrix_map
            reason = _combine_reasons(reason, matrix_reason)

        if stft_map is None:
            rejected_candidate_sheets.append(_rejection(candidate, reason or "no_stft_parser_accepted_sheet"))
            continue
        valid, validation_reason = validate_stft_grid(
            stft_map,
            target_frequency_THz=target_frequency_THz,
            search_window_THz=search_window_THz,
        )
        if not valid:
            rejected_candidate_sheets.append(_rejection(candidate, validation_reason or "stft_grid_validation_failed"))
            continue
        accepted_maps.append(stft_map)

    inspection["accepted_maps"] = [
        _map_report(stft_map, target_frequency_THz=target_frequency_THz, search_window_THz=search_window_THz)
        for stft_map in accepted_maps
    ]
    inspection["rejected_candidate_sheets"] = rejected_candidate_sheets

    if not accepted_maps:
        reason = "No sheet contained a validated time-frequency amplitude grid or long-form time/frequency/amplitude table."
        summary = _no_stft_summary(inspection, reason)
        _remove_stale_stft_outputs(output_dir)
        _write_json(output_dir / "stft_feature_summary.json", summary)
        return summary

    rows: list[dict[str, Any]] = []
    for stft_map in accepted_maps:
        rows.extend(
            extract_target_band_over_time(
                stft_map,
                target_frequency_THz=target_frequency_THz,
                search_window_THz=search_window_THz,
            )
        )
    compute_stft_recoverability(
        rows,
        target_frequency_THz=target_frequency_THz,
        search_window_THz=search_window_THz,
    )
    compute_stft_delta_persistence(rows)
    k_star_by_map = detect_stft_k_star(rows)
    group_rows = _summarize_by_map(rows, k_star_by_map)
    standout_features, interpretation, missing_notes = summarize_stft_features(
        stft_maps=accepted_maps,
        rows=rows,
        inspection=inspection,
        k_star_by_map=k_star_by_map,
        target_frequency_THz=target_frequency_THz,
    )

    _write_csv(output_dir / "stft_feature_summary.csv", rows, SUMMARY_FIELDNAMES)
    _write_csv(output_dir / "stft_recoverability_by_time.csv", group_rows, GROUP_FIELDNAMES)
    plot_notes = _write_stft_plots(rows, accepted_maps, output_dir)
    missing_notes.extend(note for note in plot_notes if note)

    recoverability_values = _finite_values(row.get("target_band_recoverability") for row in rows)
    peak_frequencies = _finite_values(row.get("peak_frequency_THz") for row in rows if row.get("peak_found"))
    k_star_values = [value for value in k_star_by_map.values() if value is not None]
    visible_failures = [row for row in rows if row.get("visible_failure")]
    target_peak_records = sum(1 for row in rows if row.get("peak_found"))
    status = "PASS"
    if rejected_candidate_sheets or target_peak_records < len(rows):
        status = "PARTIAL"

    summary = {
        "status": status,
        "target_frequency_THz": target_frequency_THz,
        "search_window_THz": search_window_THz,
        "stft_maps_found": len(accepted_maps),
        "time_points_audited": len(rows),
        "frequency_points_audited": int(sum(report.get("frequency_points", 0) for report in inspection["accepted_maps"])),
        "target_peak_records_found": target_peak_records,
        "mean_peak_frequency_THz": _mean_or_none(peak_frequencies),
        "mean_stft_recoverability": _mean_or_none(recoverability_values),
        "k_star_detected": bool(k_star_values),
        "k_star_time_ps": min(k_star_values) if k_star_values else None,
        "visible_failure_detected": bool(visible_failures),
        "bounded_interpretation": interpretation,
        "standout_features": standout_features,
        "missing_data_notes": sorted(set(missing_notes)),
        "raw_files_inspected": len(inspection.get("raw_files_inspected", [])),
        "candidate_sheets_inspected": inspection.get("candidate_sheets_inspected", 0),
        "accepted_maps": inspection.get("accepted_maps", []),
        "rejected_candidate_sheets": rejected_candidate_sheets,
        "plots": {
            "stft_target_band_recoverability": str(output_dir / "stft_target_band_recoverability.png")
            if (output_dir / "stft_target_band_recoverability.png").exists()
            else None,
            "stft_peak_frequency_over_time": str(output_dir / "stft_peak_frequency_over_time.png")
            if (output_dir / "stft_peak_frequency_over_time.png").exists()
            else None,
            "stft_amplitude_map_examples": str(output_dir / "stft_amplitude_map_examples.png")
            if (output_dir / "stft_amplitude_map_examples.png").exists()
            else None,
            "stft_delta_persistence": str(output_dir / "stft_delta_persistence.png")
            if (output_dir / "stft_delta_persistence.png").exists()
            else None,
            "stft_safety_margin": str(output_dir / "stft_safety_margin.png")
            if (output_dir / "stft_safety_margin.png").exists()
            else None,
        },
    }
    _write_json(output_dir / "stft_feature_summary.json", summary)
    return summary


def _inspect_sheet_preview(
    path: Path,
    sheet_name: str,
    preview_rows: list[tuple[Any, ...]],
    *,
    max_row: int,
    max_column: int,
    root: Path,
) -> dict[str, Any]:
    headers = list(preview_rows[0]) if preview_rows else []
    header_text = " ".join(_text(value) for value in headers)
    all_preview_text = " ".join(_text(value) for row in preview_rows for value in row)
    clean_text = _clean_key(f"{sheet_name} {header_text} {all_preview_text}")
    normalized_headers = [_normalize_header(value, index) for index, value in enumerate(headers)]
    has_time = _find_header_index(normalized_headers, TIME_ALIASES) is not None
    has_frequency = _find_header_index(normalized_headers, FREQUENCY_ALIASES) is not None
    has_amplitude = _find_header_index(normalized_headers, AMPLITUDE_ALIASES) is not None
    has_stft_keyword = any(keyword in clean_text for keyword in STFT_KEYWORDS)
    matrix_possible = _preview_matrix_possible(preview_rows)
    long_form_possible = has_time and has_frequency and has_amplitude
    candidate = bool(has_stft_keyword or long_form_possible or matrix_possible)
    if long_form_possible:
        reason = "candidate_long_form_time_frequency_amplitude_columns"
    elif matrix_possible:
        reason = "candidate_matrix_like_numeric_axes"
    elif has_stft_keyword:
        reason = "candidate_stft_keyword_present"
    elif has_time and has_amplitude and not has_frequency:
        reason = "not_candidate_time_amplitude_without_frequency_axis"
    elif has_frequency and has_amplitude and not has_time:
        reason = "not_candidate_spectrum_without_time_axis"
    else:
        reason = "not_candidate_no_time_frequency_amplitude_structure"
    return {
        "source_file": path.name,
        "relative_path": _relative_to_root(path, root),
        "sheet_name": sheet_name,
        "rows": max_row,
        "columns": max_column,
        "candidate": candidate,
        "reason": reason,
        "signals": {
            "has_time_header": has_time,
            "has_frequency_header": has_frequency,
            "has_amplitude_header": has_amplitude,
            "has_stft_keyword": has_stft_keyword,
            "matrix_like_numeric_axes": matrix_possible,
        },
        "preview_headers": [_text(value) for value in headers[:16]],
    }


def _no_stft_summary(inspection: dict[str, Any], reason: str) -> dict[str, Any]:
    rejected = inspection.get("rejected_candidate_sheets") or []
    missing_notes = list(inspection.get("missing_data_notes") or [])
    for candidate in inspection.get("candidate_sheet_reports", []):
        if not any(
            candidate.get("relative_path") == item.get("relative_path")
            and candidate.get("sheet_name") == item.get("sheet_name")
            for item in rejected
        ):
            rejected.append(_rejection(candidate, candidate.get("reason", "candidate_not_accepted")))
    for item in rejected:
        missing_notes.append(f"{item.get('source_file')}#{item.get('sheet_name')} rejected: {item.get('reason')}")
    return {
        "status": "NO_STFT_DATA_FOUND",
        "raw_files_inspected": len(inspection.get("raw_files_inspected", [])),
        "raw_file_records": inspection.get("raw_files_inspected", []),
        "candidate_sheets_inspected": inspection.get("candidate_sheets_inspected", 0),
        "candidate_sheet_reports": inspection.get("candidate_sheet_reports", []),
        "accepted_maps": [],
        "rejected_candidate_sheets": rejected,
        "reason": reason,
        "bounded_interpretation": "No raw STFT/time-frequency map was parsed. Existing spectral audit result remains unchanged.",
        "missing_data_notes": sorted(set(missing_notes)),
    }


def _map_report(
    stft_map: dict[str, Any],
    *,
    target_frequency_THz: float,
    search_window_THz: float,
) -> dict[str, Any]:
    time = np.asarray(stft_map.get("time_ps"), dtype=float)
    frequency = np.asarray(stft_map.get("frequency_THz"), dtype=float)
    finite = np.isfinite(time) & np.isfinite(frequency)
    time = time[finite]
    frequency = frequency[finite]
    return {
        "source_file": stft_map.get("source_file"),
        "sheet_name": stft_map.get("sheet_name"),
        "map_id": stft_map.get("map_id"),
        "parse_shape": stft_map.get("parse_shape"),
        "time_points": int(np.unique(time).size),
        "frequency_points": int(np.unique(frequency).size),
        "target_window_points": int(np.count_nonzero(np.abs(np.unique(frequency) - target_frequency_THz) <= search_window_THz)),
    }


def _summarize_by_map(rows: list[dict[str, Any]], k_star_by_map: dict[str, float | None]) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for map_id, map_rows in sorted(_group_rows_by_map(rows).items()):
        peak_frequencies = _finite_values(row.get("peak_frequency_THz") for row in map_rows)
        amplitudes = _finite_values(row.get("peak_amplitude") for row in map_rows)
        recoverability = _finite_values(row.get("target_band_recoverability") for row in map_rows)
        failures = [row for row in map_rows if row.get("visible_failure")]
        summaries.append(
            {
                "map_id": map_id,
                "n_time_points": len(map_rows),
                "mean_peak_frequency_THz": _mean_or_none(peak_frequencies),
                "std_peak_frequency_THz": _std_or_none(peak_frequencies),
                "mean_peak_amplitude": _mean_or_none(amplitudes),
                "mean_target_band_recoverability": _mean_or_none(recoverability),
                "min_target_band_recoverability": min(recoverability) if recoverability else None,
                "max_target_band_recoverability": max(recoverability) if recoverability else None,
                "k_star_detected": k_star_by_map.get(map_id) is not None,
                "k_star_time_ps": k_star_by_map.get(map_id),
                "visible_failure_detected": bool(failures),
                "first_visible_failure_time_ps": failures[0].get("time_ps") if failures else None,
            }
        )
    return summaries


def _write_stft_plots(rows: list[dict[str, Any]], stft_maps: list[dict[str, Any]], output_dir: Path) -> list[str]:
    notes = [
        _plot_time_metric(rows, "target_band_recoverability", output_dir / "stft_target_band_recoverability.png"),
        _plot_time_metric(rows, "peak_frequency_THz", output_dir / "stft_peak_frequency_over_time.png"),
        _plot_time_metric(rows, "delta_persistence", output_dir / "stft_delta_persistence.png"),
        _plot_time_metric(rows, "safety_margin", output_dir / "stft_safety_margin.png"),
        _plot_map_examples(stft_maps, output_dir / "stft_amplitude_map_examples.png"),
    ]
    return [note for note in notes if note]


def _plot_time_metric(rows: list[dict[str, Any]], metric: str, path: Path) -> str | None:
    series: dict[str, list[tuple[float, float]]] = {}
    for row in rows:
        time_value = _float_or_none(row.get("time_ps"))
        metric_value = _float_or_none(row.get(metric))
        if time_value is None or metric_value is None:
            continue
        series.setdefault(str(row.get("map_id", "map")), []).append((time_value, metric_value))
    if not series:
        return f"{path.name} skipped: no usable {metric} values"
    plt.figure(figsize=(7.4, 4.4))
    for label, points in sorted(series.items()):
        ordered = sorted(points)
        plt.plot([x for x, _ in ordered], [y for _, y in ordered], marker="o", linewidth=1.3, label=label[-32:])
    if metric in {"target_band_recoverability"}:
        plt.ylim(0.0, 1.05)
    plt.xlabel("time ps")
    plt.ylabel(metric)
    if len(series) <= 6:
        plt.legend(frameon=False, fontsize=7)
    plt.tight_layout()
    plt.savefig(path, dpi=160)
    plt.close()
    return None


def _plot_map_examples(stft_maps: list[dict[str, Any]], path: Path) -> str | None:
    if not stft_maps:
        return "stft_amplitude_map_examples.png skipped: no accepted STFT maps"
    stft_map = stft_maps[0]
    time = np.asarray(stft_map.get("time_ps"), dtype=float)
    frequency = np.asarray(stft_map.get("frequency_THz"), dtype=float)
    amplitude = np.asarray(stft_map.get("amplitude"), dtype=float)
    finite = np.isfinite(time) & np.isfinite(frequency) & np.isfinite(amplitude)
    time = time[finite]
    frequency = frequency[finite]
    amplitude = amplitude[finite]
    unique_time = np.unique(time)
    unique_frequency = np.unique(frequency)
    if unique_time.size < 3 or unique_frequency.size < 3:
        return "stft_amplitude_map_examples.png skipped: grid too sparse"
    grid = np.full((unique_frequency.size, unique_time.size), np.nan)
    time_index = {value: index for index, value in enumerate(unique_time)}
    frequency_index = {value: index for index, value in enumerate(unique_frequency)}
    for t_value, f_value, a_value in zip(time, frequency, amplitude):
        grid[frequency_index[f_value], time_index[t_value]] = a_value
    plt.figure(figsize=(7.4, 4.8))
    plt.imshow(
        grid,
        aspect="auto",
        origin="lower",
        extent=[float(unique_time.min()), float(unique_time.max()), float(unique_frequency.min()), float(unique_frequency.max())],
        interpolation="nearest",
    )
    plt.axhline(TARGET_FREQUENCY_THZ, color="white", linestyle="--", linewidth=1.0)
    plt.xlabel("time ps")
    plt.ylabel("frequency THz")
    plt.colorbar(label="amplitude")
    plt.tight_layout()
    plt.savefig(path, dpi=160)
    plt.close()
    return None


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _csv_value(row.get(field)) for field in fieldnames})


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _remove_stale_stft_outputs(output_dir: Path) -> None:
    for name in STALE_STFT_OUTPUTS:
        path = output_dir / name
        if path.exists():
            path.unlink()


def _rejection(candidate: dict[str, Any], reason: Any) -> dict[str, Any]:
    return {
        "source_file": candidate.get("source_file"),
        "relative_path": candidate.get("relative_path"),
        "sheet_name": candidate.get("sheet_name"),
        "reason": str(reason),
        "signals": candidate.get("signals", {}),
        "preview_headers": candidate.get("preview_headers", []),
    }


def _combine_reasons(first: str | None, second: str | None) -> str:
    if first and second:
        return f"{first}; {second}"
    return first or second or "no_parser_reason_available"


def _group_rows_by_map(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(str(row.get("map_id", "unknown_map")), []).append(row)
    return grouped


def _visible_failure(row: dict[str, Any]) -> bool:
    if not row.get("peak_found"):
        return True
    amplitude_retention = _float_or_none(row.get("amplitude_retention"))
    recoverability = _float_or_none(row.get("target_band_recoverability"))
    if amplitude_retention is not None and amplitude_retention < 0.40:
        return True
    if recoverability is not None and recoverability < 0.40:
        return True
    return False


def _confidence_from_background(peak_to_background: float | None, recoverability: Any) -> float | None:
    p2b = _float_or_none(peak_to_background)
    if p2b is not None:
        return _clip01(p2b / 10.0)
    return _clip01(_float_or_none(recoverability))


def _find_header_row(raw_rows: list[tuple[Any, ...]]) -> tuple[int | None, list[Any] | None]:
    for index, raw_row in enumerate(raw_rows[:12]):
        values = list(raw_row)
        text_cells = [value for value in values if value is not None and not _looks_float(value)]
        if len(text_cells) >= 2:
            return index, values
    return None, None


def _find_header_index(headers: list[str], aliases: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if _clean_key(header) in aliases:
            return index
    for index, header in enumerate(headers):
        clean = _clean_key(header)
        if any(alias != "thz" and alias in clean for alias in aliases):
            return index
    return None


def _normalize_header(value: Any, index: int) -> str:
    text = _text(value)
    if not text:
        return f"col_{index}"
    normalized = text.strip().lower()
    normalized = normalized.replace("frequency (thz)", "frequency_THz")
    normalized = normalized.replace("freq. (thz)", "frequency_THz")
    normalized = normalized.replace("time (ps)", "time_ps")
    normalized = normalized.replace("delay (ps)", "delay_ps")
    normalized = normalized.replace("dt_p (ps)", "dt_p_ps")
    normalized = normalized.replace("dt (ps)", "dt_ps")
    normalized = re.sub(r"[^a-zA-Z0-9]+", "_", normalized).strip("_")
    aliases = {
        "frequency_THz": "frequency_THz",
        "frequency_thz": "frequency_THz",
        "freq_thz": "frequency_THz",
        "thz": "frequency_THz",
        "time_ps": "time_ps",
        "delay_ps": "delay_ps",
        "dt_ps": "time_ps",
        "dt_p_ps": "time_ps",
        "thz_signal": "amplitude",
        "stft_int": "stft_amplitude",
        "stft_intensity": "stft_amplitude",
        "amplitude": "amplitude",
        "intensity": "amplitude",
        "power": "amplitude",
    }
    return aliases.get(normalized, normalized or f"col_{index}")


def _preview_matrix_possible(preview_rows: list[tuple[Any, ...]]) -> bool:
    if len(preview_rows) < 4:
        return False
    header = list(preview_rows[0])
    numeric_header = sum(1 for value in header[1:] if _float_or_none(value) is not None)
    numeric_first_column = 0
    for row in preview_rows[1:]:
        if _float_or_none(_cell(row, 0)) is not None:
            numeric_first_column += 1
    numeric_body = 0
    for row in preview_rows[1:]:
        for value in list(row)[1:8]:
            if _float_or_none(value) is not None:
                numeric_body += 1
    return numeric_header >= 3 and numeric_first_column >= 3 and numeric_body >= 9


def _axis_includes_target(values: list[float], target_frequency_THz: float, search_window_THz: float) -> bool:
    return any(abs(value - target_frequency_THz) <= search_window_THz for value in values)


def _relative_to_root(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def _safe_id(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "stft_map"


def _cell(row: tuple[Any, ...] | list[Any], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _text(value).lower())


def _looks_float(value: Any) -> bool:
    return _float_or_none(value) is not None


def _float_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def _clip01(value: float | None) -> float | None:
    if value is None or not math.isfinite(float(value)):
        return None
    return min(max(float(value), 0.0), 1.0)


def _finite_values(values: Any) -> list[float]:
    finite: list[float] = []
    for value in values:
        numeric = _float_or_none(value)
        if numeric is not None:
            finite.append(numeric)
    return finite


def _mean_or_none(values: list[float]) -> float | None:
    if not values:
        return None
    return float(sum(values) / len(values))


def _std_or_none(values: list[float]) -> float | None:
    if not values:
        return None
    mean = sum(values) / len(values)
    return float(math.sqrt(sum((value - mean) ** 2 for value in values) / len(values)))


def _csv_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return f"{value:.10g}"
    return value
