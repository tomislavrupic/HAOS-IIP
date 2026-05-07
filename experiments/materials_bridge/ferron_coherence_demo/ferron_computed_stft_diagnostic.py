#!/usr/bin/env python3
"""Computed STFT diagnostic from raw Fig. 2 time traces.

This module derives a time-frequency diagnostic from the downloaded raw
transient-reflectance traces. It does not reinterpret the XLSX files as raw
published STFT maps. Outputs are explicitly labeled
``computed_from_time_trace_stft_diagnostic``.
"""

from __future__ import annotations

import csv
import json
import math
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np


ROOT = Path(__file__).resolve().parent
OUTPUTS_DIR = ROOT / "outputs"
RAW_FIG2 = OUTPUTS_DIR / "raw" / "Source data Fig.2.xlsx"

TARGET_FREQUENCY_THZ = 3.13
FREQUENCY_TOLERANCE_THZ = 0.05
WINDOW_PS = 10.0
STEP_PS = 5.0
TRACE_SHEET_THICKNESS_NM = {"Fig.2a": 224.0, "Fig.2b": 96.0}
TRACE_SHEET_THICKNESS_ASSIGNMENT_NOTE = (
    "Fig.2a/Fig.2b headers do not include thickness labels; assignment follows "
    "the published Fig.2d propagation peak-time ordering."
)

TRACE_FIELDNAMES = [
    "source_file",
    "sheet_name",
    "trace_id",
    "distance_um",
    "thickness_nm",
    "time_ps",
    "computed_target_band_intensity",
    "normalized_intensity",
]

PEAK_FIELDNAMES = [
    "source_file",
    "sheet_name",
    "trace_id",
    "distance_um",
    "thickness_nm",
    "n_time_points",
    "peak_time_ps",
    "peak_intensity",
    "fwhm_width_ps",
    "arrival_delay_from_zero_um_ps",
]

VELOCITY_FIELDNAMES = [
    "source_file",
    "sheet_name",
    "thickness_nm",
    "n_distances",
    "slope_ps_per_um",
    "intercept_ps",
    "group_velocity_m_s",
    "r_squared",
    "monotonic_peak_delay",
]

COMPARISON_FIELDNAMES = [
    "computed_trace_id",
    "published_trace_id",
    "distance_um",
    "thickness_nm",
    "n_overlap_points",
    "computed_peak_time_ps",
    "published_peak_time_ps",
    "peak_time_delta_ps",
    "envelope_correlation",
    "mean_abs_envelope_error",
]


def run_computed_stft_diagnostic(
    *,
    root: Path = ROOT,
    output_dir: Path = OUTPUTS_DIR,
    target_frequency_THz: float = TARGET_FREQUENCY_THZ,
    frequency_tolerance_THz: float = FREQUENCY_TOLERANCE_THZ,
    window_ps: float = WINDOW_PS,
    step_ps: float = STEP_PS,
) -> dict[str, Any]:
    """Run the derived STFT diagnostic and write CSV/JSON/plots."""

    diagnostic_dir = output_dir / "computed_stft_diagnostic"
    diagnostic_dir.mkdir(parents=True, exist_ok=True)
    raw_path = root / "outputs" / "raw" / "Source data Fig.2.xlsx"
    raw_traces, notes = _load_raw_fig2_time_traces(raw_path, root)
    if not raw_traces:
        summary = {
            "status": "NO_RAW_TIME_TRACE_DATA",
            "mode": "computed_from_time_trace_stft_diagnostic",
            "raw_stft_time_frequency_grid_status": "NO_STFT_DATA_FOUND",
            "traces_computed": 0,
            "bounded_interpretation": "No computed STFT diagnostic was run because raw Fig. 2 time traces were unavailable.",
            "missing_data_notes": notes,
        }
        _write_json(diagnostic_dir / "computed_stft_summary.json", summary)
        _write_csv(diagnostic_dir / "computed_stft_trace_records.csv", [], TRACE_FIELDNAMES)
        _write_csv(diagnostic_dir / "computed_stft_peak_times.csv", [], PEAK_FIELDNAMES)
        _write_csv(diagnostic_dir / "computed_stft_group_velocity.csv", [], VELOCITY_FIELDNAMES)
        _write_csv(diagnostic_dir / "computed_vs_published_stft_comparison.csv", [], COMPARISON_FIELDNAMES)
        return summary

    trace_records: list[dict[str, Any]] = []
    peak_rows: list[dict[str, Any]] = []
    for raw_trace in raw_traces:
        computed = compute_target_band_intensity(
            np.asarray(raw_trace["time_ps"], dtype=float),
            np.asarray(raw_trace["signal"], dtype=float),
            target_frequency_THz=target_frequency_THz,
            frequency_tolerance_THz=frequency_tolerance_THz,
            window_ps=window_ps,
            step_ps=step_ps,
        )
        if computed is None:
            notes.append(f"{raw_trace['trace_id']} skipped: STFT target band could not be computed")
            continue
        times, intensity = computed
        max_intensity = float(np.nanmax(intensity)) if intensity.size else 0.0
        if max_intensity <= 1.0e-18:
            notes.append(f"{raw_trace['trace_id']} skipped: computed target-band intensity is zero")
            continue
        normalized = intensity / max_intensity
        for time_value, value, norm_value in zip(times, intensity, normalized):
            trace_records.append(
                {
                    "source_file": raw_trace["source_file"],
                    "sheet_name": raw_trace["sheet_name"],
                    "trace_id": raw_trace["trace_id"],
                    "distance_um": raw_trace["distance_um"],
                    "thickness_nm": raw_trace["thickness_nm"],
                    "time_ps": float(time_value),
                    "computed_target_band_intensity": float(value),
                    "normalized_intensity": float(norm_value),
                }
            )
        peak_rows.append(
            _summarize_trace_peak(
                raw_trace,
                times,
                normalized,
            )
        )

    velocity_rows = _estimate_group_velocity(peak_rows)
    comparison_rows = _compare_computed_to_published(trace_records, output_dir, root)
    plot_notes = _write_plots(trace_records, peak_rows, velocity_rows, comparison_rows, diagnostic_dir)
    _write_csv(diagnostic_dir / "computed_stft_trace_records.csv", trace_records, TRACE_FIELDNAMES)
    _write_csv(diagnostic_dir / "computed_stft_peak_times.csv", peak_rows, PEAK_FIELDNAMES)
    _write_csv(diagnostic_dir / "computed_stft_group_velocity.csv", velocity_rows, VELOCITY_FIELDNAMES)
    _write_csv(diagnostic_dir / "computed_vs_published_stft_comparison.csv", comparison_rows, COMPARISON_FIELDNAMES)

    velocities = _finite_values(row.get("group_velocity_m_s") for row in velocity_rows)
    correlations = _finite_values(row.get("envelope_correlation") for row in comparison_rows)
    peak_deltas = _finite_values(abs(row.get("peak_time_delta_ps")) for row in comparison_rows)
    interpretation = (
        "Computed STFT target-band envelopes were derived from the raw Fig. 2 time traces "
        "using a 10 ps window and 5 ps step. This is a derived diagnostic, not raw published "
        "STFT-grid parsing. It can be compared to the published Fig. 2d target-band STFT "
        "intensity traces as a bounded reproducibility check."
    )
    summary = {
        "status": "PASS" if trace_records else "NO_COMPUTED_STFT_RECORDS",
        "mode": "computed_from_time_trace_stft_diagnostic",
        "raw_stft_time_frequency_grid_status": "NO_STFT_DATA_FOUND",
        "source_sheets": sorted(TRACE_SHEET_THICKNESS_NM),
        "sheet_thickness_assignment_nm": TRACE_SHEET_THICKNESS_NM,
        "sheet_thickness_assignment_note": TRACE_SHEET_THICKNESS_ASSIGNMENT_NOTE,
        "target_frequency_THz": target_frequency_THz,
        "frequency_tolerance_THz": frequency_tolerance_THz,
        "window_ps": window_ps,
        "step_ps": step_ps,
        "traces_computed": len({row["trace_id"] for row in trace_records}),
        "time_points_computed": len(trace_records),
        "peak_records": len(peak_rows),
        "velocity_groups": len(velocity_rows),
        "mean_group_velocity_m_s": _mean_or_none(velocities),
        "comparison_records": len(comparison_rows),
        "mean_envelope_correlation_vs_published": _mean_or_none(correlations),
        "mean_abs_peak_time_delta_vs_published_ps": _mean_or_none(peak_deltas),
        "group_velocity_rows": velocity_rows,
        "bounded_interpretation": interpretation,
        "missing_data_notes": sorted(set(notes + plot_notes)),
        "outputs_dir": str(diagnostic_dir),
    }
    _write_json(diagnostic_dir / "computed_stft_summary.json", summary)
    return summary


def compute_target_band_intensity(
    time_ps: np.ndarray,
    signal: np.ndarray,
    *,
    target_frequency_THz: float,
    frequency_tolerance_THz: float,
    window_ps: float,
    step_ps: float,
) -> tuple[np.ndarray, np.ndarray] | None:
    """Compute a NumPy STFT and extract target-band intensity."""

    time = np.asarray(time_ps, dtype=float)
    values = np.asarray(signal, dtype=float)
    finite = np.isfinite(time) & np.isfinite(values)
    time = time[finite]
    values = values[finite]
    if time.size < 16:
        return None
    order = np.argsort(time)
    time = time[order]
    values = values[order] - float(np.nanmean(values[order]))
    dt = float(np.nanmedian(np.diff(time)))
    if not math.isfinite(dt) or dt <= 0.0:
        return None
    window_size = max(int(round(window_ps / dt)), 4)
    hop_size = max(int(round(step_ps / dt)), 1)
    if values.size < window_size:
        return None
    window = np.hanning(window_size)
    frequency = np.fft.rfftfreq(window_size, d=dt)
    mask = np.abs(frequency - target_frequency_THz) <= frequency_tolerance_THz
    if not np.any(mask):
        return None
    centers: list[float] = []
    intensities: list[float] = []
    for start in range(0, values.size - window_size + 1, hop_size):
        stop = start + window_size
        segment = values[start:stop]
        segment = segment - float(np.nanmean(segment))
        spectrum = np.fft.rfft(segment * window)
        power = np.abs(spectrum) ** 2
        centers.append(float(np.nanmean(time[start:stop])))
        intensities.append(float(np.nanmean(power[mask])))
    if not centers:
        return None
    return np.asarray(centers, dtype=float), np.asarray(intensities, dtype=float)


def _load_raw_fig2_time_traces(path: Path, root: Path) -> tuple[list[dict[str, Any]], list[str]]:
    notes: list[str] = []
    traces: list[dict[str, Any]] = []
    if not path.exists():
        return [], [f"Missing raw workbook: {path}"]
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        return [], [f"openpyxl unavailable for raw Fig. 2 parsing: {exc}"]

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name, thickness_nm in TRACE_SHEET_THICKNESS_NM.items():
            if sheet_name not in workbook.sheetnames:
                notes.append(f"{path.name}#{sheet_name} skipped: sheet missing")
                continue
            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            if not rows:
                notes.append(f"{path.name}#{sheet_name} skipped: sheet empty")
                continue
            headers = list(rows[0])
            for time_index in range(0, len(headers) - 1, 2):
                signal_index = time_index + 1
                label = str(headers[signal_index] or "")
                distance = _distance_from_label(label)
                if distance is None:
                    notes.append(f"{path.name}#{sheet_name}:{label} skipped: missing distance")
                    continue
                time_values: list[float] = []
                signal_values: list[float] = []
                for row in rows[1:]:
                    time_value = _float_or_none(_cell(row, time_index))
                    signal_value = _float_or_none(_cell(row, signal_index))
                    if time_value is None or signal_value is None:
                        continue
                    time_values.append(time_value)
                    signal_values.append(signal_value)
                if len(time_values) < 16:
                    notes.append(f"{path.name}#{sheet_name}:{label} skipped: too few points")
                    continue
                traces.append(
                    {
                        "source_file": _relative_to_root(path, root),
                        "sheet_name": sheet_name,
                        "trace_id": _safe_id(f"{path.stem}_{sheet_name}_{label}"),
                        "distance_um": distance,
                        "thickness_nm": thickness_nm,
                        "time_ps": np.asarray(time_values, dtype=float),
                        "signal": np.asarray(signal_values, dtype=float),
                    }
                )
    finally:
        workbook.close()
    return traces, notes


def _load_published_trace_records(output_dir: Path, root: Path) -> list[dict[str, Any]]:
    csv_path = output_dir / "published_stft_trace_records.csv"
    if csv_path.exists():
        rows: list[dict[str, Any]] = []
        with csv_path.open("r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                rows.append(row)
        return rows
    # Fallback only reads the raw workbook; it does not create raw STFT-grid claims.
    from stft_feature_audit import _parse_published_stft_intensity_traces  # type: ignore[attr-defined]

    rows, _ = _parse_published_stft_intensity_traces(root)
    return rows


def _summarize_trace_peak(raw_trace: dict[str, Any], times: np.ndarray, values: np.ndarray) -> dict[str, Any]:
    peak_index = int(np.nanargmax(values))
    width, _ = _fwhm_width_from_trace(times, values, peak_index)
    return {
        "source_file": raw_trace["source_file"],
        "sheet_name": raw_trace["sheet_name"],
        "trace_id": raw_trace["trace_id"],
        "distance_um": raw_trace["distance_um"],
        "thickness_nm": raw_trace["thickness_nm"],
        "n_time_points": int(times.size),
        "peak_time_ps": float(times[peak_index]),
        "peak_intensity": float(values[peak_index]),
        "fwhm_width_ps": width,
        "arrival_delay_from_zero_um_ps": None,
    }


def _estimate_group_velocity(peak_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, float], list[dict[str, Any]]] = {}
    for row in peak_rows:
        thickness = _float_or_none(row.get("thickness_nm"))
        if thickness is None:
            continue
        key = (str(row.get("source_file")), str(row.get("sheet_name")), thickness)
        grouped.setdefault(key, []).append(row)

    velocity_rows: list[dict[str, Any]] = []
    for (source_file, sheet_name, thickness), rows in sorted(grouped.items()):
        valid = [
            (distance, peak_time, row)
            for row in rows
            if (distance := _float_or_none(row.get("distance_um"))) is not None
            and (peak_time := _float_or_none(row.get("peak_time_ps"))) is not None
        ]
        valid.sort(key=lambda item: item[0])
        if len(valid) < 2:
            continue
        baseline = valid[0][1]
        for _, peak_time, row in valid:
            row["arrival_delay_from_zero_um_ps"] = peak_time - baseline
        distances = np.asarray([item[0] for item in valid], dtype=float)
        peak_times = np.asarray([item[1] for item in valid], dtype=float)
        slope, intercept = np.polyfit(distances, peak_times, 1)
        predicted = intercept + slope * distances
        residual = float(np.sum((peak_times - predicted) ** 2))
        total = float(np.sum((peak_times - float(np.mean(peak_times))) ** 2))
        velocity = 1.0e6 / float(slope) if slope > 1.0e-12 else None
        velocity_rows.append(
            {
                "source_file": source_file,
                "sheet_name": sheet_name,
                "thickness_nm": thickness,
                "n_distances": len(valid),
                "slope_ps_per_um": float(slope),
                "intercept_ps": float(intercept),
                "group_velocity_m_s": velocity,
                "r_squared": 1.0 - residual / total if total > 1.0e-12 else None,
                "monotonic_peak_delay": _monotonic_increase_fraction([item[1] for item in valid]),
            }
        )
    return velocity_rows


def _compare_computed_to_published(
    computed_records: list[dict[str, Any]],
    output_dir: Path,
    root: Path,
) -> list[dict[str, Any]]:
    published_records = _load_published_trace_records(output_dir, root)
    computed_by_key = _group_trace_records(computed_records)
    published_by_key = _group_trace_records(published_records)
    rows: list[dict[str, Any]] = []
    for key, computed in sorted(computed_by_key.items()):
        published = published_by_key.get(key)
        if not published:
            continue
        computed_ordered = sorted(computed, key=lambda row: _float_or_none(row.get("time_ps")) or 0.0)
        published_ordered = sorted(published, key=lambda row: _float_or_none(row.get("time_ps")) or 0.0)
        c_time = np.asarray([_float_or_none(row.get("time_ps")) for row in computed_ordered], dtype=float)
        c_value = np.asarray([_float_or_none(row.get("normalized_intensity")) for row in computed_ordered], dtype=float)
        p_time = np.asarray([_float_or_none(row.get("time_ps")) for row in published_ordered], dtype=float)
        p_value = np.asarray([_float_or_none(row.get("normalized_intensity")) for row in published_ordered], dtype=float)
        finite_c = np.isfinite(c_time) & np.isfinite(c_value)
        finite_p = np.isfinite(p_time) & np.isfinite(p_value)
        c_time = c_time[finite_c]
        c_value = c_value[finite_c]
        p_time = p_time[finite_p]
        p_value = p_value[finite_p]
        if c_time.size < 3 or p_time.size < 3:
            continue
        overlap = (p_time >= float(np.min(c_time))) & (p_time <= float(np.max(c_time)))
        if np.count_nonzero(overlap) < 3:
            continue
        p_overlap_time = p_time[overlap]
        p_overlap_value = p_value[overlap]
        c_interp = np.interp(p_overlap_time, c_time, c_value)
        corr = _correlation(c_interp, p_overlap_value)
        error = float(np.mean(np.abs(c_interp - p_overlap_value)))
        computed_peak_time = float(c_time[int(np.nanargmax(c_value))])
        published_peak_time = float(p_time[int(np.nanargmax(p_value))])
        first = computed_ordered[0]
        rows.append(
            {
                "computed_trace_id": first.get("trace_id"),
                "published_trace_id": published_ordered[0].get("trace_id"),
                "distance_um": _float_or_none(first.get("distance_um")),
                "thickness_nm": _float_or_none(first.get("thickness_nm")),
                "n_overlap_points": int(np.count_nonzero(overlap)),
                "computed_peak_time_ps": computed_peak_time,
                "published_peak_time_ps": published_peak_time,
                "peak_time_delta_ps": computed_peak_time - published_peak_time,
                "envelope_correlation": corr,
                "mean_abs_envelope_error": error,
            }
        )
    return rows


def _write_plots(
    trace_records: list[dict[str, Any]],
    peak_rows: list[dict[str, Any]],
    velocity_rows: list[dict[str, Any]],
    comparison_rows: list[dict[str, Any]],
    output_dir: Path,
) -> list[str]:
    notes = []
    for note in (
        _plot_trace_records(trace_records, output_dir / "computed_stft_intensity_traces.png"),
        _plot_peak_time_vs_distance(peak_rows, output_dir / "computed_stft_peak_time_vs_distance.png"),
        _plot_velocity_rows(velocity_rows, output_dir / "computed_stft_group_velocity.png"),
        _plot_comparison_rows(comparison_rows, output_dir / "computed_vs_published_stft_comparison.png"),
    ):
        if note:
            notes.append(note)
    return notes


def _plot_trace_records(records: list[dict[str, Any]], path: Path) -> str | None:
    grouped = _group_by_trace_id(records)
    if not grouped:
        return f"{path.name} skipped: no computed STFT traces"
    plt.figure(figsize=(8.0, 4.8))
    for _, rows in sorted(grouped.items()):
        ordered = sorted(rows, key=lambda row: _float_or_none(row.get("time_ps")) or 0.0)
        label = f"{_format_number(ordered[0].get('thickness_nm'))} nm, {_format_number(ordered[0].get('distance_um'))} um"
        plt.plot(
            [_float_or_none(row.get("time_ps")) for row in ordered],
            [_float_or_none(row.get("normalized_intensity")) for row in ordered],
            linewidth=1.2,
            label=label,
        )
    plt.xlabel("computed STFT time ps")
    plt.ylabel("normalized target-band intensity")
    plt.ylim(0.0, 1.05)
    plt.legend(frameon=False, fontsize=7, ncol=2)
    plt.tight_layout()
    plt.savefig(path, dpi=160)
    plt.close()
    return None


def _plot_peak_time_vs_distance(rows: list[dict[str, Any]], path: Path) -> str | None:
    series = _series_by_thickness(rows, "peak_time_ps")
    if not series:
        return f"{path.name} skipped: no peak rows"
    plt.figure(figsize=(7.2, 4.4))
    for label, points in sorted(series.items()):
        ordered = sorted(points)
        plt.plot([x for x, _ in ordered], [y for _, y in ordered], marker="o", linewidth=1.4, label=label)
    plt.xlabel("distance um")
    plt.ylabel("computed STFT peak time ps")
    plt.legend(frameon=False, fontsize=8)
    plt.tight_layout()
    plt.savefig(path, dpi=160)
    plt.close()
    return None


def _plot_velocity_rows(rows: list[dict[str, Any]], path: Path) -> str | None:
    points = [
        (thickness, velocity)
        for row in rows
        if (thickness := _float_or_none(row.get("thickness_nm"))) is not None
        and (velocity := _float_or_none(row.get("group_velocity_m_s"))) is not None
    ]
    if not points:
        return f"{path.name} skipped: no computed group velocity values"
    ordered = sorted(points)
    plt.figure(figsize=(6.6, 4.2))
    plt.bar([str(_format_number(x)) for x, _ in ordered], [y for _, y in ordered])
    plt.axhline(100000.0, color="black", linestyle="--", linewidth=1.0, label="1e5 m/s reference")
    plt.xlabel("thickness nm")
    plt.ylabel("computed group velocity m/s")
    plt.legend(frameon=False, fontsize=8)
    plt.tight_layout()
    plt.savefig(path, dpi=160)
    plt.close()
    return None


def _plot_comparison_rows(rows: list[dict[str, Any]], path: Path) -> str | None:
    points = [
        (distance, delta, thickness)
        for row in rows
        if (distance := _float_or_none(row.get("distance_um"))) is not None
        and (delta := _float_or_none(row.get("peak_time_delta_ps"))) is not None
        and (thickness := _float_or_none(row.get("thickness_nm"))) is not None
    ]
    if not points:
        return f"{path.name} skipped: no comparison rows"
    plt.figure(figsize=(7.2, 4.4))
    for thickness in sorted(set(point[2] for point in points)):
        subset = sorted((distance, delta) for distance, delta, item_thickness in points if item_thickness == thickness)
        plt.plot(
            [x for x, _ in subset],
            [y for _, y in subset],
            marker="o",
            linewidth=1.4,
            label=f"{_format_number(thickness)} nm",
        )
    plt.axhline(0.0, color="black", linestyle="--", linewidth=1.0)
    plt.xlabel("distance um")
    plt.ylabel("computed minus published peak time ps")
    plt.legend(frameon=False, fontsize=8)
    plt.tight_layout()
    plt.savefig(path, dpi=160)
    plt.close()
    return None


def _group_trace_records(records: list[dict[str, Any]]) -> dict[tuple[float, float], list[dict[str, Any]]]:
    grouped: dict[tuple[float, float], list[dict[str, Any]]] = {}
    for row in records:
        distance = _float_or_none(row.get("distance_um"))
        thickness = _float_or_none(row.get("thickness_nm"))
        if distance is None or thickness is None:
            continue
        grouped.setdefault((distance, thickness), []).append(row)
    return grouped


def _group_by_trace_id(records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in records:
        grouped.setdefault(str(row.get("trace_id")), []).append(row)
    return grouped


def _series_by_thickness(rows: list[dict[str, Any]], metric: str) -> dict[str, list[tuple[float, float]]]:
    series: dict[str, list[tuple[float, float]]] = {}
    for row in rows:
        distance = _float_or_none(row.get("distance_um"))
        value = _float_or_none(row.get(metric))
        thickness = _float_or_none(row.get("thickness_nm"))
        if distance is None or value is None or thickness is None:
            continue
        series.setdefault(f"{_format_number(thickness)} nm", []).append((distance, value))
    return series


def _fwhm_width_from_trace(times: np.ndarray, values: np.ndarray, peak_index: int) -> tuple[float | None, str | None]:
    peak = float(values[peak_index])
    baseline = float(np.nanmin(values))
    height = peak - baseline
    if height <= 1.0e-12:
        return None, "peak_height_not_above_trace_baseline"
    half = baseline + 0.5 * height
    left = _trace_half_crossing(times, values, peak_index, half, direction=-1)
    right = _trace_half_crossing(times, values, peak_index, half, direction=1)
    if left is None:
        left = float(times[0])
    if right is None:
        right = float(times[-1])
    return float(right - left), None


def _trace_half_crossing(
    times: np.ndarray,
    values: np.ndarray,
    peak_index: int,
    half: float,
    *,
    direction: int,
) -> float | None:
    index = peak_index
    while 0 <= index + direction < values.size:
        next_index = index + direction
        if values[next_index] <= half:
            x0 = float(times[index])
            x1 = float(times[next_index])
            y0 = float(values[index])
            y1 = float(values[next_index])
            if abs(y1 - y0) <= 1.0e-12:
                return x1
            fraction = (half - y0) / (y1 - y0)
            return x0 + fraction * (x1 - x0)
        index = next_index
    return None


def _distance_from_label(label: str) -> float | None:
    import re

    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*u?m\b", label, flags=re.IGNORECASE)
    if not match:
        return None
    return float(match.group(1))


def _monotonic_increase_fraction(values: list[float]) -> float | None:
    if len(values) < 2:
        return None
    return sum(1 for left, right in zip(values, values[1:]) if right >= left) / (len(values) - 1)


def _correlation(left: np.ndarray, right: np.ndarray) -> float | None:
    if left.size < 2 or right.size != left.size:
        return None
    left_std = float(np.std(left))
    right_std = float(np.std(right))
    if left_std <= 1.0e-12 or right_std <= 1.0e-12:
        return None
    return float(np.corrcoef(left, right)[0, 1])


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _csv_value(row.get(field)) for field in fieldnames})


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _cell(row: Any, index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _relative_to_root(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return str(path)


def _safe_id(value: str) -> str:
    import re

    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "computed_stft_trace"


def _float_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(numeric):
        return None
    return numeric


def _finite_values(values: Any) -> list[float]:
    finite: list[float] = []
    for value in values:
        numeric = _float_or_none(value)
        if numeric is not None:
            finite.append(numeric)
    return finite


def _mean_or_none(values: list[float]) -> float | None:
    if not values:
        return None
    return float(sum(values) / len(values))


def _csv_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return f"{value:.10g}"
    return value


def _format_number(value: Any) -> str:
    numeric = _float_or_none(value)
    if numeric is None:
        return "unknown"
    if abs(numeric - round(numeric)) <= 1.0e-9:
        return str(int(round(numeric)))
    return f"{numeric:.6g}"


if __name__ == "__main__":
    run_computed_stft_diagnostic()
